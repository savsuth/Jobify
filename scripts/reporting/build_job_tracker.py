"""Builds output/job_tracker.xlsx from the current DB state.

Read-only against the DB. Sheets: CURRENT TARGETS (currently eligible jobs
only), JOBS (full history), REVIEW (ambiguous eligibility), DUPLICATES,
RESUMES, SUMMARY.

Usage: python scripts/reporting/build_job_tracker.py [--baseline-max-job-id N]

--baseline-max-job-id marks jobs with id > N as "NEW THIS RUN" in CURRENT
TARGETS, everything else "EXISTING ELIGIBLE". Omit it for a plain rebuild.
"""

import argparse
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from src.discovery.normalize import (
    CURRENTLY_ELIGIBLE,
    AMBIGUOUS_REQUIRES_REVIEW,
    classify_location,
    classify_seniority,
    is_target_role,
)
from src.reporting.excel_style import (
    FONT_NAME,
    HEADER_FILL,
    HEADER_FONT,
    BASE_FONT,
    add_hyperlink,
    add_table,
    as_naive_dt,
    autosize,
    style_header,
    write_row,
)
from src.reporting.tracker import fetch_records, resolve_representatives

OUT_PATH = Path(__file__).resolve().parents[2] / "output" / "job_tracker.xlsx"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--baseline-max-job-id", type=int, default=None)
    args = parser.parse_args()

    records, db_counts, posted_within_days = fetch_records()
    assert len(records) == db_counts["jobs"]

    missing_tex = [r["id"] for r in records if r["draft_id"] is not None and r["resume_tex_path"] and not Path(r["resume_tex_path"]).exists()]
    missing_pdf = [r["id"] for r in records if r["draft_id"] is not None and r["resume_link"] is None]
    if missing_tex:
        raise RuntimeError(
            f"{len(missing_tex)} resume draft(s) have no .tex artifact file at all - run "
            f"scripts/backfill/backfill_resume_artifacts.py first: {missing_tex}"
        )
    if missing_pdf:
        print(f"WARNING: {len(missing_pdf)} resume draft(s) have a .tex but no PDF yet "
              f"(shown as Pending in Excel, not treated as a build failure): {missing_pdf}")

    representative_ids, representative_reason = resolve_representatives(records)
    by_canonical: dict[str, list[dict]] = {}
    for rec in records:
        by_canonical.setdefault(rec["canonical_key"], []).append(rec)
    duplicate_groups = {ck: recs for ck, recs in by_canonical.items() if len(recs) > 1}

    jobs_sheet_records = sorted([r for r in records if r["id"] in representative_ids], key=lambda r: r["id"])
    current_targets = sorted(
        [r for r in jobs_sheet_records if r["current_eligibility"] == CURRENTLY_ELIGIBLE], key=lambda r: r["id"]
    )
    review_records = sorted(
        [r for r in records if r["current_eligibility"] == AMBIGUOUS_REQUIRES_REVIEW], key=lambda r: r["id"]
    )

    # Re-check CURRENT TARGETS against the actual rules rather than trusting
    # current_eligibility alone - catches drift if the eligibility rules
    # change without a full re-backfill.
    violations = []
    for rec in current_targets:
        loc = classify_location(rec["location"])
        sen = classify_seniority(rec["title"], rec["description_raw"] or "")
        role_ok = is_target_role(rec["title"])
        if loc == "non_us" or sen == "senior" or not role_ok:
            violations.append((rec["id"], loc, sen, role_ok))
    if violations:
        raise RuntimeError(f"CURRENT TARGETS contains {len(violations)} known rule violation(s): {violations}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # CURRENT TARGETS sheet
    ws_ct = wb.create_sheet("CURRENT TARGETS")
    ct_headers = [
        "Job ID", "Company", "Job Title", "Location", "Remote Type", "Employment Type",
        "Job URL", "Source", "Source ID", "Canonical Key", "Date Discovered", "Discovery State",
        "ATS Score", "ATS Status", "Resume Status", "Resume Type", "Resume Draft ID",
        "Resume File Name", "Resume Link", "Resume Created", "Processing Status", "Notes",
        "Posted Date", "Freshness Status", "Days Since Posted",
    ]
    ws_ct.append(ct_headers)
    for rec in current_targets:
        note = ""
        if rec["canonical_key"] in duplicate_groups:
            other_ids = [x["id"] for x in duplicate_groups[rec["canonical_key"]] if x["id"] != rec["id"]]
            note = f"Duplicate posting also discovered as job ID(s) {other_ids} - see DUPLICATES sheet"
        resume_status = rec["status"] if rec["status"] in ("RESUME NOT REQUIRED", "RESUME DECLINED", "RESUME COMPLETED") else "Pending"
        if args.baseline_max_job_id is not None:
            discovery_state = "NEW THIS RUN" if rec["id"] > args.baseline_max_job_id else "EXISTING ELIGIBLE"
        else:
            discovery_state = "N/A"
        r = write_row(ws_ct, [
            rec["id"], rec["company"], rec["title"], rec["location"], rec["remote_type"],
            rec["employment_type"], rec["url"], rec["source"], rec["source_id"], rec["canonical_key"],
            as_naive_dt(rec["discovered_at"]), discovery_state,
            rec["ats_score"], rec["ats_status"] or "Pending", resume_status,
            rec["resume_source"] or "Pending", rec["draft_id"] if rec["draft_id"] is not None else "Pending",
            rec["resume_file_name"] or "Pending", None,
            as_naive_dt(rec["draft_created_at"]) if rec["draft_created_at"] else "Pending",
            rec["status"], note,
            as_naive_dt(rec["posted_at"]) if rec["posted_at"] else "Unknown", rec["freshness_status"],
            rec["days_since_posted"] if rec["days_since_posted"] is not None else "Unknown",
        ], wrap_cols=(3, 22), date_cols=(11, 20, 23), fill_key=rec["status"], fill_col=21)
        add_hyperlink(ws_ct, r, 7, rec["url"], rec["url"])
        if rec["resume_link"]:
            add_hyperlink(ws_ct, r, 19, rec["resume_link"], rec["resume_file_name"])
    style_header(ws_ct, len(ct_headers))
    autosize(ws_ct, [8, 16, 34, 16, 12, 14, 38, 11, 22, 34, 16, 15, 10, 13, 15, 13, 13, 26, 30, 16, 19, 44, 16, 18, 14])
    add_table(ws_ct, "CurrentTargetsTable", len(ct_headers), len(current_targets))
    ws_ct.row_dimensions[1].height = 30

    # JOBS sheet - full history
    ws_j = wb.create_sheet("JOBS")
    j_headers = [
        "Job ID", "Company", "Job Title", "Location", "Remote Type", "Employment Type",
        "Job URL", "Source", "Source ID", "Canonical Key", "Date Discovered",
        "Current Eligibility", "Eligibility Reason", "ATS Score", "ATS Status",
        "Resume Status", "Resume Type", "Resume Draft ID", "Resume File Name",
        "Resume Link", "Resume Created", "Processing Status", "Notes",
        "Posted Date", "Freshness Status", "Days Since Posted",
    ]
    ws_j.append(j_headers)
    for rec in jobs_sheet_records:
        note = ""
        if rec["canonical_key"] in duplicate_groups:
            other_ids = [x["id"] for x in duplicate_groups[rec["canonical_key"]] if x["id"] != rec["id"]]
            note = f"Duplicate posting also discovered as job ID(s) {other_ids} - see DUPLICATES sheet"
        resume_status = rec["status"] if rec["status"] in ("RESUME NOT REQUIRED", "RESUME DECLINED", "RESUME COMPLETED") else "Pending"
        r = write_row(ws_j, [
            rec["id"], rec["company"], rec["title"], rec["location"], rec["remote_type"],
            rec["employment_type"], rec["url"], rec["source"], rec["source_id"], rec["canonical_key"],
            as_naive_dt(rec["discovered_at"]), rec["current_eligibility"],
            "; ".join(rec["eligibility_reasons"]) or "N/A",
            rec["ats_score"], rec["ats_status"] or "Pending", resume_status,
            rec["resume_source"] or "Pending", rec["draft_id"] if rec["draft_id"] is not None else "Pending",
            rec["resume_file_name"] or "Pending", None,
            as_naive_dt(rec["draft_created_at"]) if rec["draft_created_at"] else "Pending",
            rec["status"], note,
            as_naive_dt(rec["posted_at"]) if rec["posted_at"] else "Unknown", rec["freshness_status"],
            rec["days_since_posted"] if rec["days_since_posted"] is not None else "Unknown",
        ], wrap_cols=(3, 13, 23), date_cols=(11, 21, 24), fill_key=rec["current_eligibility"], fill_col=12)
        add_hyperlink(ws_j, r, 7, rec["url"], rec["url"])
        if rec["resume_link"]:
            add_hyperlink(ws_j, r, 20, rec["resume_link"], rec["resume_file_name"])
    style_header(ws_j, len(j_headers))
    autosize(ws_j, [8, 16, 34, 16, 12, 14, 38, 11, 22, 34, 16, 22, 40, 10, 13, 15, 13, 13, 26, 30, 16, 19, 44, 16, 18, 14])
    add_table(ws_j, "JobsTable", len(j_headers), len(jobs_sheet_records))
    ws_j.row_dimensions[1].height = 30

    # REVIEW sheet - ambiguous eligibility
    ws_rv = wb.create_sheet("REVIEW")
    rv_headers = ["Job ID", "Company", "Job Title", "Location", "Source", "Job URL",
                  "Eligibility Reason", "Raw Location Evidence", "ATS Status", "Notes"]
    ws_rv.append(rv_headers)
    for rec in review_records:
        r = write_row(ws_rv, [
            rec["id"], rec["company"], rec["title"], rec["location"], rec["source"], rec["url"],
            "; ".join(rec["eligibility_reasons"]), rec["location"] or "N/A", rec["ats_status"] or "Pending",
            "Requires human review before eligible for ATS/Resume processing",
        ], wrap_cols=(3, 7, 10))
        add_hyperlink(ws_rv, r, 6, rec["url"], rec["url"])
    style_header(ws_rv, len(rv_headers))
    autosize(ws_rv, [8, 16, 34, 16, 11, 40, 40, 20, 13, 44])
    add_table(ws_rv, "ReviewTable", len(rv_headers), len(review_records))

    # DUPLICATES sheet
    ws_d = wb.create_sheet("DUPLICATES")
    d_headers = ["Canonical Key", "Job IDs", "Company", "Titles", "Sources", "URLs",
                 "ATS IDs / Statuses", "Resume IDs / Statuses", "Representative ID", "Representative Reason"]
    ws_d.append(d_headers)
    for ck in sorted(duplicate_groups.keys()):
        recs = sorted(duplicate_groups[ck], key=lambda r: r["id"])
        rep_id, rep_reason = representative_reason[ck]
        job_ids = ", ".join(str(r["id"]) for r in recs)
        companies = ", ".join(sorted(set(r["company"] for r in recs)))
        titles = " | ".join(f"[{r['id']}] {r['title']}" for r in recs)
        sources = ", ".join(f"{r['id']}={r['source']}" for r in recs)
        urls = " | ".join(f"[{r['id']}] {r['url']}" for r in recs)
        ats = ", ".join(f"{r['id']}={r['ats_score']}/{r['ats_status']}" for r in recs)
        resumes = ", ".join(
            f"{r['id']}={r['resume_source'] or 'Pending'}" + (f"(draft {r['draft_id']})" if r["draft_id"] else "")
            for r in recs
        )
        write_row(ws_d, [ck, job_ids, companies, titles, sources, urls, ats, resumes, rep_id, rep_reason],
                  wrap_cols=(1, 4, 6, 7, 8))
    style_header(ws_d, len(d_headers))
    autosize(ws_d, [40, 14, 14, 50, 22, 55, 30, 30, 14, 28])
    add_table(ws_d, "DuplicatesTable", len(d_headers), len(duplicate_groups))

    # RESUMES sheet
    ws_r = wb.create_sheet("RESUMES")
    r_headers = ["Job ID", "Company", "Job Title", "ATS Score", "Resume Type", "Tailored vs Master",
                 "Resume Selection", "Draft ID", "File Name", "File Path", "Resume Link",
                 "Resume Created", "Status", "Project Selected", "Reason / Notes"]
    ws_r.append(r_headers)
    resume_records = sorted([r for r in records if r["resume_source"] is not None], key=lambda r: r["id"])
    for rec in resume_records:
        tailored_vs_master = "Tailored" if rec["tailoring_needed"] else "Master"
        projects = ""
        if rec["project_selection"] and isinstance(rec["project_selection"], dict):
            sel = rec["project_selection"].get("selected_projects") or []
            projects = ", ".join(p.get("name", "") for p in sel if isinstance(p, dict))
        r = write_row(ws_r, [
            rec["id"], rec["company"], rec["title"], rec["ats_score"], rec["resume_source"],
            tailored_vs_master, rec["resume_source"], rec["draft_id"] if rec["draft_id"] is not None else "Pending",
            rec["resume_file_name"] or "N/A", rec["resume_file_path"] or "N/A", None,
            as_naive_dt(rec["draft_created_at"]) if rec["draft_created_at"] else "Pending",
            rec["status"], projects or "N/A", (rec["summary_of_changes"] or rec["rs_reasoning"] or ""),
        ], wrap_cols=(10, 14, 15), date_cols=(12,))
        if rec["resume_link"]:
            add_hyperlink(ws_r, r, 11, rec["resume_link"], rec["resume_file_name"])
    style_header(ws_r, len(r_headers))
    autosize(ws_r, [8, 16, 34, 10, 13, 15, 15, 9, 26, 46, 30, 16, 22, 30, 60])
    add_table(ws_r, "ResumesTable", len(r_headers), len(resume_records))

    # SUMMARY sheet
    ws_s = wb.create_sheet("SUMMARY")
    ws_s.column_dimensions["A"].width = 40
    ws_s.column_dimensions["B"].width = 14
    ws_s.cell(row=1, column=1, value="Job Application Tracker - Summary").font = Font(name=FONT_NAME, bold=True, size=14)
    ws_s.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = Font(name=FONT_NAME, italic=True, size=9)
    if args.baseline_max_job_id is not None:
        ws_s.cell(row=3, column=1, value=f"Run identifier: baseline_max_job_id={args.baseline_max_job_id} (jobs with id > this value are NEW THIS RUN)").font = Font(name=FONT_NAME, italic=True, size=9)

    n_j, n_ct, n_res = len(jobs_sheet_records), len(current_targets), len(resume_records)
    last_j, last_ct, last_res = n_j + 1, n_ct + 1, n_res + 1
    n_new_this_run = sum(1 for r in current_targets if args.baseline_max_job_id is not None and r["id"] > args.baseline_max_job_id)

    metrics = [
        ("Total historical jobs (DB)", db_counts["jobs"]),
        ("Total jobs on JOBS sheet (canonical, deduped)", f"=COUNTA(JOBS!A2:A{last_j})"),
        ("CURRENT TARGETS (currently eligible)", f"=COUNTA('CURRENT TARGETS'!A2:A{last_ct})"),
        ("CURRENT TARGETS - NEW THIS RUN", n_new_this_run if args.baseline_max_job_id is not None else "N/A"),
        ("Currently ineligible (JOBS sheet)", f'=COUNTIF(JOBS!L2:L{last_j},"CURRENTLY_INELIGIBLE")'),
        ("Ambiguous / requires review (JOBS sheet)", f'=COUNTIF(JOBS!L2:L{last_j},"AMBIGUOUS_REQUIRES_REVIEW")'),
        ("ATS analyzed", f'=COUNTIFS(JOBS!O2:O{last_j},"passed")+COUNTIFS(JOBS!O2:O{last_j},"rejected")+COUNTIFS(JOBS!O2:O{last_j},"ineligible")'),
        ("ATS passed", f'=COUNTIF(JOBS!O2:O{last_j},"passed")'),
        ("ATS rejected", f'=COUNTIF(JOBS!O2:O{last_j},"rejected")'),
        ("Resumes completed", f'=COUNTIF(JOBS!V2:V{last_j},"RESUME COMPLETED")'),
        ("Resumes pending", f'=COUNTIF(JOBS!V2:V{last_j},"ATS PASSED — RESUME PENDING")'),
        ("Resumes declined", f'=COUNTIF(JOBS!V2:V{last_j},"RESUME DECLINED")'),
        ("Resume artifacts available (RESUMES sheet)", f'=COUNTIF(RESUMES!H2:H{last_res},">0")'),
        ("Duplicate groups", len(duplicate_groups)),
        ("Companies represented (JOBS sheet)", f"=SUMPRODUCT(1/COUNTIF(JOBS!B2:B{last_j},JOBS!B2:B{last_j}))"),
        ("Sources represented (JOBS sheet)", f"=SUMPRODUCT(1/COUNTIF(JOBS!H2:H{last_j},JOBS!H2:H{last_j}))"),
        ("Freshness gate (posted_within_days)", posted_within_days if posted_within_days is not None else "disabled (null)"),
        ("POSTED_RECENTLY (JOBS sheet)", f'=COUNTIF(JOBS!Y2:Y{last_j},"POSTED_RECENTLY")'),
        ("POSTED_OLD (JOBS sheet)", f'=COUNTIF(JOBS!Y2:Y{last_j},"POSTED_OLD")'),
        ("POST_DATE_UNKNOWN (JOBS sheet)", f'=COUNTIF(JOBS!Y2:Y{last_j},"POST_DATE_UNKNOWN")'),
        ("CURRENT TARGETS - POSTED_RECENTLY", f'=COUNTIF(\'CURRENT TARGETS\'!X2:X{last_ct},"POSTED_RECENTLY")'),
    ]
    r = 4
    ws_s.cell(row=r, column=1, value="Metric").font = HEADER_FONT
    ws_s.cell(row=r, column=2, value="Value").font = HEADER_FONT
    ws_s.cell(row=r, column=1).fill = HEADER_FILL
    ws_s.cell(row=r, column=2).fill = HEADER_FILL
    r += 1
    for label, val in metrics:
        ws_s.cell(row=r, column=1, value=label).font = BASE_FONT
        ws_s.cell(row=r, column=2, value=val).font = BASE_FONT
        r += 1

    for section_title, items, formula_col in [
        ("Ineligibility / ambiguity reason breakdown (JOBS sheet, current eligibility)",
         sorted(_count_reasons(jobs_sheet_records).items(), key=lambda x: -x[1]), None),
        ("Company breakdown (JOBS sheet)", sorted(_count_by(jobs_sheet_records, "company").items(), key=lambda x: -x[1]), "B"),
        ("Source breakdown (JOBS sheet)", sorted(_count_by(jobs_sheet_records, "source").items(), key=lambda x: -x[1]), "H"),
    ]:
        r += 1
        ws_s.cell(row=r, column=1, value=section_title).font = Font(name=FONT_NAME, bold=True)
        r += 1
        ws_s.cell(row=r, column=1, value="Item").font = HEADER_FONT
        ws_s.cell(row=r, column=2, value="Count").font = HEADER_FONT
        ws_s.cell(row=r, column=1).fill = HEADER_FILL
        ws_s.cell(row=r, column=2).fill = HEADER_FILL
        r += 1
        for label, count in items:
            ws_s.cell(row=r, column=1, value=label).font = BASE_FONT
            if formula_col:
                ws_s.cell(row=r, column=2, value=f'=COUNTIF(JOBS!{formula_col}2:{formula_col}{last_j},A{r})').font = BASE_FONT
            else:
                ws_s.cell(row=r, column=2, value=count).font = BASE_FONT
            r += 1

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Saved workbook to {OUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"CURRENT TARGETS: {n_ct} rows | JOBS: {n_j} rows | DUPLICATES: {len(duplicate_groups)} rows | "
          f"RESUMES: {n_res} rows | REVIEW: {len(review_records)} rows")
    print(f"DB counts used: {db_counts}")


def _count_reasons(records):
    counts: dict[str, int] = {}
    for rec in records:
        for reason in rec["eligibility_reasons"]:
            counts[reason] = counts.get(reason, 0) + 1
    return counts


def _count_by(records, key):
    counts: dict[str, int] = {}
    for rec in records:
        counts[rec[key]] = counts.get(rec[key], 0) + 1
    return counts


if __name__ == "__main__":
    main()
