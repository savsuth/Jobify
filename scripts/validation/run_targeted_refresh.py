"""One-off historical run: a bounded, diversity-capped Discovery -> ATS ->
Resume Agent cycle, kept here as a record of that run rather than as a
script meant to be re-run as-is.

Calls the real production functions directly and adds nothing to
Discovery/ATS/Resume Agent itself - only a local selection layer on top:
an extra title exclusion, a stricter citizenship/clearance bar, a rough
defense/government classification, and a priority ranking with
per-company and Palantir caps. Writes to the same tables a normal
pipeline run would, and appends a "NEW JOBS" sheet to job_tracker.xlsx
without touching its other sheets.

Makes real Greenhouse/Lever/web-search and Claude API calls.
"""

import re
from pathlib import Path

from langgraph.graph import END
from langgraph.types import Send

from src.db.models import CandidateProfile, Job, SearchConfig
from src.db.session import get_session_factory
from src.discovery.aggregate import _filter as base_filter
from src.discovery.greenhouse import fetch_greenhouse_jobs
from src.discovery.lever import fetch_lever_jobs
from src.discovery.normalize import (
    CURRENTLY_ELIGIBLE,
    NormalizedJob,
    canonical_job_identity,
    classify_location,
    classify_seniority,
    compute_current_eligibility,
    is_target_role,
)
from src.discovery.web_search import fetch_web_search_jobs
from src.graph.pipeline import _persist, analyze_job, resume_agent_node, select_master_resume
from src.integrations.github import fetch_github_context
from src.reporting.excel_style import BASE_FONT, BORDER, HEADER_FILL, HEADER_FONT, LINK_FONT, WRAP
from src.reporting.tracker import RESUME_COMPLETED, RESUME_DECLINED, RESUME_NOT_REQUIRED, fetch_records

import httpx
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT_PATH = Path(__file__).resolve().parents[2] / "output" / "job_tracker.xlsx"

MAX_NEW_JOBS = 12
MAX_PER_COMPANY = 2
MAX_PALANTIR = 2

# Run-scoped additional title exclusion (on top of the permanent
# senior/staff/principal/lead check). Title-only, same precision reasoning
# as the permanent classifiers - never scans description prose.
_HARD_EXCLUDE_TITLE_PATTERN = re.compile(
    r"\b(senior|sr\.?|staff|principal|lead|manager|director|head|distinguished)\b", re.IGNORECASE
)

_DEFENSE_PATTERN = re.compile(
    r"\b(defense|defence|department of defense|dod|national security|intelligence community|"
    r"federal government|public sector|government contract|military|homeland security|"
    r"classified (?:information|program|systems)|intelligence agenc(?:y|ies)|counterterrorism)\b",
    re.IGNORECASE,
)

_ROLE_TIERS = [
    (1, re.compile(r"\bbackend\b.{0,20}\bengineer\b", re.IGNORECASE)),
    (2, re.compile(r"\bsoftware engineer\b|\bsde\s*(?:ii|2)\b", re.IGNORECASE)),
    (3, re.compile(r"\b(?:api|platform|infrastructure)\b.{0,20}\bengineer\b", re.IGNORECASE)),
    (4, re.compile(r"\b(?:machine learning|ml) engineer\b", re.IGNORECASE)),
    (5, re.compile(r"\bai engineer\b|\bapplied ai engineer\b", re.IGNORECASE)),
    (6, re.compile(r"\bdata scientist\b|\banalytics engineer\b", re.IGNORECASE)),
    (7, re.compile(r"\binfrastructure engineer\b|\bplatform engineer\b", re.IGNORECASE)),
    (8, re.compile(r"\bforward deployed\b", re.IGNORECASE)),
]
_EXPERIENCE_RANK = {"entry_level": 0, "0-1": 1, "1-2": 2, "unknown": 3, "2-3": 4, "3-5": 5, "5+": 6}


def role_priority_tier(title: str) -> int:
    for tier, pattern in _ROLE_TIERS:
        if pattern.search(title):
            return tier
    return 9


def classify_defense(title: str, description: str) -> bool:
    return bool(_DEFENSE_PATTERN.search(f"{title}\n{description}"))


def citizenship_clearance_exclusion(work_auth_status: str, work_auth_reason: str | None) -> tuple[bool, str | None]:
    """Reads the already-computed work_auth_status/work_auth_reason off the
    NormalizedJob - no new classification logic. A stricter bar for THIS
    curated batch only: "ineligible" is always excluded; "unknown" is ALSO
    excluded here when the reason mentions clearance/classified/"US Person"
    (this run wanted zero clearance-mentioning jobs) - ordinary ambiguous
    sponsorship-only language stays included, per the explicit instruction
    that "no sponsorship"/"authorized to work" wording is NOT a citizenship
    requirement."""
    reason_l = (work_auth_reason or "").lower()
    if work_auth_status == "ineligible":
        bucket = "clearance" if "clearance" in reason_l else "citizenship"
        return True, bucket
    if work_auth_status == "unknown" and any(m in reason_l for m in ("clearance", "classified", "us person")):
        bucket = "clearance" if "clearance" in reason_l else "citizenship"
        return True, bucket
    return False, None


def _new_filter_stats() -> dict:
    return {
        "raw_fetched": 0, "non_us_rejected": 0, "non_full_time_rejected": 0,
        "title_rejected": 0, "senior_rejected": 0, "eligible": 0,
    }


def fetch_per_company(config: SearchConfig) -> tuple[dict[str, list[NormalizedJob]], dict[str, int], dict[str, dict]]:
    """Mirrors src.discovery.aggregate.discover_candidates() exactly (same
    fetch_greenhouse_jobs/fetch_lever_jobs/fetch_web_search_jobs/_filter calls,
    same live sources, no new fetching logic) - restructured only to keep
    PER-COMPANY raw/filtered counts (aggregate.discover_candidates aggregates
    stats globally), needed for the Palantir-specific and per-source reporting
    this refresh requires."""
    candidates_by_company: dict[str, list[NormalizedJob]] = {}
    raw_counts: dict[str, int] = {}
    stats_by_company: dict[str, dict] = {}

    with httpx.Client(timeout=30) as client:
        for token in config.companies.get("greenhouse", []):
            key = f"greenhouse:{token}"
            raw_jobs = fetch_greenhouse_jobs(token, client=client)
            stats = _new_filter_stats()
            candidates_by_company[key] = base_filter(raw_jobs, stats)
            raw_counts[key] = len(raw_jobs)
            stats_by_company[key] = stats
        for token in config.companies.get("lever", []):
            key = f"lever:{token}"
            raw_jobs = fetch_lever_jobs(token, client=client)
            stats = _new_filter_stats()
            candidates_by_company[key] = base_filter(raw_jobs, stats)
            raw_counts[key] = len(raw_jobs)
            stats_by_company[key] = stats

    web_jobs = fetch_web_search_jobs({
        "title_keywords": config.title_keywords,
        "locations": config.locations,
        "remote_pref": config.remote_pref,
        "experience_level": config.experience_level,
    })
    stats = _new_filter_stats()
    candidates_by_company["web:search"] = base_filter(web_jobs, stats)
    raw_counts["web:search"] = len(web_jobs)
    stats_by_company["web:search"] = stats

    return candidates_by_company, raw_counts, stats_by_company


def main() -> None:
    session = get_session_factory()()
    try:
        config = session.query(SearchConfig).first()
        if config is None:
            raise RuntimeError("search_config is empty - run scripts/maintenance/seed_config.py first")

        baseline_max_job_id = session.query(Job.id).order_by(Job.id.desc()).first()
        baseline_max_job_id = baseline_max_job_id[0] if baseline_max_job_id else 0
        print(f"Baseline max job id (pre-run): {baseline_max_job_id}")

        print("\n=== LIVE DISCOVERY (real Greenhouse + Lever + web-search calls) ===")
        candidates_by_company, raw_counts, stats_by_company = fetch_per_company(config)

        total_raw = sum(raw_counts.values())
        total_stats = _new_filter_stats()
        for s in stats_by_company.values():
            for k in total_stats:
                total_stats[k] += s[k]
        print(f"Raw fetched: {total_raw} | Greenhouse: {sum(v for k, v in raw_counts.items() if k.startswith('greenhouse:'))} "
              f"| Lever: {sum(v for k, v in raw_counts.items() if k.startswith('lever:'))} | Web search: {raw_counts.get('web:search', 0)}")
        print(f"Base-filter stats: {total_stats}")

        # --- dedup against DB (source,source_id) + canonical_key - identical
        # logic to discover_jobs() in src/graph/pipeline.py, not modified there,
        # just re-run here read-only against the live DB state.
        existing_keys = {(s, sid) for s, sid in session.query(Job.source, Job.source_id).all()}
        existing_canonical_keys = {ck for (ck,) in session.query(Job.canonical_key).all()}
        seen_canonical_this_run: set[str] = set()

        funnel = []  # one entry per base-filter-surviving candidate, full audit trail
        for company_key, jobs in candidates_by_company.items():
            for nj in jobs:
                ck = canonical_job_identity(nj.url)
                entry = {"nj": nj, "company_key": company_key, "canonical_key": ck}
                if (nj.source, nj.source_id) in existing_keys:
                    entry["dedup_status"] = "dup_db_source_id"
                elif ck in existing_canonical_keys or ck in seen_canonical_this_run:
                    entry["dedup_status"] = "dup_canonical"
                else:
                    entry["dedup_status"] = "new"
                    seen_canonical_this_run.add(ck)
                funnel.append(entry)

        genuinely_new = [f for f in funnel if f["dedup_status"] == "new"]
        dup_count = len(funnel) - len(genuinely_new)
        print(f"Genuinely new (post source_id + canonical dedup): {len(genuinely_new)} | duplicates removed: {dup_count}")

        # --- CURRENT ELIGIBILITY gate - same function used in production
        # (_persist/_pending_jobs), re-run here read-only before persistence.
        for f in genuinely_new:
            nj = f["nj"]
            status, reasons = compute_current_eligibility(nj.title, nj.location, nj.description_raw)
            f["current_eligibility"] = status
            f["eligibility_reasons"] = reasons

        currently_eligible = [f for f in genuinely_new if f["current_eligibility"] == CURRENTLY_ELIGIBLE]
        rejected_non_us = sum(1 for f in genuinely_new if "non-US location" in f["eligibility_reasons"])
        rejected_wrong_role = sum(1 for f in genuinely_new if "wrong role family" in f["eligibility_reasons"])
        rejected_senior_permanent = sum(1 for f in genuinely_new if "senior/staff/principal/lead title" in f["eligibility_reasons"])
        rejected_non_full_time = sum(1 for f in genuinely_new if any(r.startswith("non-full-time") for r in f["eligibility_reasons"]))
        ambiguous_count = sum(1 for f in genuinely_new if f["current_eligibility"] == "AMBIGUOUS_REQUIRES_REVIEW")

        # --- ADDITIONAL run-level restrictions (this script only - see module
        # docstring): extra title exclusion, citizenship/clearance, defense class.
        additional_title_excluded = 0
        citizenship_excluded = 0
        clearance_excluded = 0
        for f in currently_eligible:
            nj = f["nj"]
            hard_title = bool(_HARD_EXCLUDE_TITLE_PATTERN.search(nj.title))
            cc_excluded, cc_bucket = citizenship_clearance_exclusion(nj.work_auth_status, nj.work_auth_reason)
            defense = classify_defense(nj.title, nj.description_raw)
            f["hard_title_excluded"] = hard_title
            f["citizenship_clearance_excluded"] = cc_excluded
            f["citizenship_clearance_bucket"] = cc_bucket
            f["defense"] = defense
            f["defense_class"] = ("A" if (defense and cc_excluded) else ("B" if defense else "C"))
            f["excluded"] = hard_title or cc_excluded
            if hard_title:
                additional_title_excluded += 1
            if cc_excluded:
                if cc_bucket == "citizenship":
                    citizenship_excluded += 1
                else:
                    clearance_excluded += 1

        eligible_pool = [f for f in currently_eligible if not f["excluded"]]
        defense_a = sum(1 for f in currently_eligible if f["defense_class"] == "A")
        defense_b = sum(1 for f in eligible_pool if f["defense_class"] == "B")
        defense_c = sum(1 for f in eligible_pool if f["defense_class"] == "C")

        print(f"\nCURRENT ELIGIBILITY: eligible={len(currently_eligible)} | rejected non-US={rejected_non_us} "
              f"| rejected senior={rejected_senior_permanent} | rejected wrong-role={rejected_wrong_role} "
              f"| rejected non-full-time={rejected_non_full_time} | ambiguous={ambiguous_count}")
        print(f"ADDITIONAL RESTRICTIONS: extra title-excluded={additional_title_excluded} "
              f"| citizenship-restricted={citizenship_excluded} | clearance-restricted={clearance_excluded} "
              f"| defense A(restricted)={defense_a} B(open)={defense_b} C(non-defense)={defense_c}")
        print(f"Genuinely new, fully-eligible pool (pre diversity cap): {len(eligible_pool)}")

        # --- Palantir-specific funnel report (section 5) ---
        palantir_funnel = [f for f in funnel if f["company_key"] == "lever:palantir"]
        palantir_new = [f for f in palantir_funnel if f["dedup_status"] == "new"]
        palantir_eligible = [f for f in palantir_new if f.get("current_eligibility") == CURRENTLY_ELIGIBLE]
        palantir_cc_excluded = sum(1 for f in palantir_eligible if f.get("citizenship_clearance_excluded"))
        palantir_defense_excluded = sum(1 for f in palantir_eligible if f.get("defense_class") == "A")
        palantir_survivors = [f for f in palantir_eligible if not f.get("excluded")]

        # --- priority ranking + diversity/Palantir cap (run-scoped selection,
        # not a permanent Discovery change - see module docstring) ---
        def sort_key(f):
            nj = f["nj"]
            tier = role_priority_tier(nj.title)
            defense_rank = 0 if f["defense_class"] == "C" else 1
            exp_rank = _EXPERIENCE_RANK.get(classify_seniority(nj.title, nj.description_raw or ""), 3)
            return (tier, defense_rank, exp_rank, nj.company, nj.title)

        eligible_pool_sorted = sorted(eligible_pool, key=sort_key)

        selected: list[dict] = []
        company_counts: dict[str, int] = {}
        palantir_selected = 0
        for f in eligible_pool_sorted:
            company_key = f["nj"].company.strip().lower()
            is_palantir = "palantir" in company_key
            if is_palantir and palantir_selected >= MAX_PALANTIR:
                continue
            if company_counts.get(company_key, 0) >= MAX_PER_COMPANY:
                continue
            selected.append(f)
            company_counts[company_key] = company_counts.get(company_key, 0) + 1
            if is_palantir:
                palantir_selected += 1
            if len(selected) >= MAX_NEW_JOBS:
                break

        palantir_excluded_by_cap = max(0, len(palantir_survivors) - palantir_selected)

        print(f"\nDIVERSITY: companies represented in selection={len(company_counts)} | max/company={MAX_PER_COMPANY}")
        print(f"PALANTIR: raw_fetched={raw_counts.get('lever:palantir', 0)} | eligible(current)={len(palantir_eligible)} "
              f"| excluded citizenship/clearance={palantir_cc_excluded} | excluded defense(A)={palantir_defense_excluded} "
              f"| retained={palantir_selected} | excluded by 2-job cap={palantir_excluded_by_cap}")
        print(f"Selected for ATS (bounded, priority+diversity ranked): {len(selected)} (cap {MAX_NEW_JOBS})")

        assert len(selected) <= MAX_NEW_JOBS, "selection exceeded the hard batch cap"

        why_selected_by_key: dict[str, str] = {}
        for f in selected:
            nj = f["nj"]
            tier = role_priority_tier(nj.title)
            why_selected_by_key[f["canonical_key"]] = (
                f"Role-priority tier {tier}; defense class {f['defense_class']}; "
                f"seniority={classify_seniority(nj.title, nj.description_raw or '')}; "
                f"company diversity slot {company_counts.get(nj.company.strip().lower(), '?')}/{MAX_PER_COMPANY}"
            )

        # --- persist via the REAL, unmodified _persist() (same canonical_key +
        # current_eligibility computation as every normal pipeline run) ---
        persisted = _persist(session, [f["nj"] for f in selected])
        persisted_ids = [p["id"] for p in persisted]
        print(f"\nPersisted {len(persisted)} new job row(s): ids {persisted_ids}")
    finally:
        session.close()

    if not persisted:
        print("\nNo genuinely-new, fully-eligible jobs survived this run's filters/caps - nothing to process.")
        _write_new_jobs_sheet([], baseline_max_job_id, {})
        print("\nFINAL VERDICT: FAIL - LIVE REFRESH QUALITY ISSUE FOUND (no eligible jobs this run)")
        return

    # --- ATS (REAL Claude calls, reusing pipeline.analyze_job() unmodified) ---
    session2 = get_session_factory()()
    try:
        profile_row = session2.query(CandidateProfile).first()
        if profile_row is None:
            raise RuntimeError("candidate_profile is empty - profile extraction hasn't run yet")
        profile_json = profile_row.structured_json
        master_raw_latex = profile_row.raw_latex
    finally:
        session2.close()

    print("\n=== ATS (real Claude calls, reusing src.graph.pipeline.analyze_job unmodified) ===")
    github_context = fetch_github_context(master_raw_latex)

    ats_calls_made = 0
    ats_results = {}  # job_id -> (status, score)
    resumes_generated = 0
    resumes_master_only = 0
    for job in persisted:
        if ats_calls_made >= MAX_NEW_JOBS:
            print("Hit absolute ATS call cap - stopping.")
            break
        state = {
            "job": job,
            "profile": profile_json,
            "reject_threshold": config.ats_threshold,
            "no_tailor_threshold": config.resume_no_tailor_threshold,
            "master_raw_latex": master_raw_latex,
            "github_context": github_context,
        }
        command = analyze_job(state)
        ats_calls_made += 1
        results = command.update.get("results", [])
        if not results:
            print(f"  job {job['id']} ({job['company']} - {job['title']}): ATS call failed, left pending.")
            continue
        r = results[0]
        ats_results[job["id"]] = (r["status"], r["match_score"])
        print(f"  job {job['id']} ({job['company']} - {job['title']}): score={r['match_score']} status={r['status']}")

        if command.goto == END:
            continue
        if isinstance(command.goto, Send):
            if command.goto.node == "select_master_resume":
                select_master_resume(command.goto.arg)
                resumes_master_only += 1
            elif command.goto.node == "resume_agent":
                resume_agent_node(command.goto.arg)
                resumes_generated += 1

    print(f"\nATS calls made: {ats_calls_made} | passed->master: {resumes_master_only} | passed->resume_agent: {resumes_generated}")

    print("\n=== Independent re-verification against DB + filesystem ===")
    records, db_counts, _posted_within_days = fetch_records()
    new_records = [r for r in records if r["id"] > baseline_max_job_id]
    print(f"DB counts now: {db_counts}")
    print(f"New job rows this run (re-read from DB): {len(new_records)}")

    _write_new_jobs_sheet(new_records, baseline_max_job_id, why_selected_by_key, funnel_by_canonical={
        f["canonical_key"]: f for f in selected
    })

    _final_report(
        total_raw=total_raw, total_stats=total_stats, genuinely_new=len(genuinely_new), dup_count=dup_count,
        currently_eligible=len(currently_eligible), rejected_non_us=rejected_non_us,
        rejected_senior=rejected_senior_permanent, rejected_wrong_role=rejected_wrong_role,
        rejected_non_full_time=rejected_non_full_time, citizenship_excluded=citizenship_excluded,
        clearance_excluded=clearance_excluded, defense_a=defense_a, defense_b=defense_b, defense_c=defense_c,
        company_counts=company_counts, palantir_raw=raw_counts.get("lever:palantir", 0),
        palantir_eligible=len(palantir_eligible), palantir_cc_excluded=palantir_cc_excluded,
        palantir_defense_excluded=palantir_defense_excluded, palantir_retained=palantir_selected,
        palantir_cap_excluded=palantir_excluded_by_cap, ats_calls_made=ats_calls_made,
        resumes_master_only=resumes_master_only, resumes_generated=resumes_generated,
        new_records=new_records, baseline_max_job_id=baseline_max_job_id, db_counts=db_counts,
    )


def _quality_check(records: list[dict]) -> tuple[list[dict], list[tuple]]:
    """Section 14: programmatic re-verification of every NEW JOBS row,
    independent of the selection logic above - re-derives each check from the
    row's own persisted fields, never trusts in-memory selection bookkeeping."""
    clean, violations = [], []
    for r in records:
        loc = classify_location(r["location"])
        role_ok = is_target_role(r["title"])
        sen = classify_seniority(r["title"], r["description_raw"] or "")
        hard_title = bool(_HARD_EXCLUDE_TITLE_PATTERN.search(r["title"]))
        emp_ok = r["employment_type"] in ("full_time", "unknown")
        problems = []
        if loc == "non_us":
            problems.append("non-US location")
        if sen == "senior" or hard_title:
            problems.append("senior/staff/principal/lead/manager/director/head/distinguished title")
        if not role_ok:
            problems.append("wrong role family")
        if not emp_ok:
            problems.append("non-full-time employment")
        if r["current_eligibility"] != CURRENTLY_ELIGIBLE:
            problems.append(f"current_eligibility={r['current_eligibility']}")
        if r["id"] <= 0 or not r["canonical_key"]:
            problems.append("invalid canonical_key")
        if problems:
            violations.append((r["id"], problems))
        else:
            clean.append(r)
    return clean, violations


def _write_new_jobs_sheet(new_records, baseline_max_job_id, why_selected_by_key, funnel_by_canonical=None):
    funnel_by_canonical = funnel_by_canonical or {}
    clean_records, violations = _quality_check(new_records) if new_records else ([], [])
    if violations:
        print(f"\nQUALITY CHECK: removed {len(violations)} row(s) that failed programmatic re-verification: {violations}")

    def _resume_readiness_rank(status):
        order = {RESUME_COMPLETED: 0, RESUME_NOT_REQUIRED: 1, RESUME_DECLINED: 2}
        return order.get(status, 3)

    def _final_sort_key(r):
        score = r["ats_score"] if r["ats_score"] is not None else -1
        fentry = funnel_by_canonical.get(r["canonical_key"])
        defense_rank = 0 if (fentry and fentry.get("defense_class") == "C") else 1
        tier = role_priority_tier(r["title"])
        return (-score, defense_rank, tier, _resume_readiness_rank(r["status"]), r["id"])

    clean_records.sort(key=_final_sort_key)

    if OUT_PATH.exists():
        wb = openpyxl.load_workbook(OUT_PATH)
    else:
        raise RuntimeError(f"{OUT_PATH} does not exist - run scripts/reporting/build_job_tracker.py first")

    if "NEW JOBS" in wb.sheetnames:
        del wb["NEW JOBS"]
    ws = wb.create_sheet("NEW JOBS")

    headers = [
        "Job ID", "Company", "Job Title", "Location", "Remote Type", "Employment Type",
        "Job URL", "Source", "Canonical Key", "Discovery Date", "ATS Score", "ATS Status",
        "Resume Status", "Resume Type", "Resume PDF File Name", "Resume PDF Link",
        "Resume Created", "Citizenship Restriction", "Security Clearance Restriction",
        "Defense Classification", "Why Selected", "Notes",
    ]
    ws.append(headers)

    for r in clean_records:
        fentry = funnel_by_canonical.get(r["canonical_key"])
        defense_class = fentry["defense_class"] if fentry else "C"
        resume_status = r["status"] if r["status"] in (RESUME_NOT_REQUIRED, RESUME_DECLINED, RESUME_COMPLETED) else "Pending"
        why = why_selected_by_key.get(r["canonical_key"], "")
        notes = "; ".join(r["eligibility_reasons"]) if r["eligibility_reasons"] else "No known rule violations"

        row_values = [
            r["id"], r["company"], r["title"], r["location"], r["remote_type"], r["employment_type"],
            r["url"], r["source"], r["canonical_key"], None,
            r["ats_score"], r["ats_status"] or "Pending", resume_status, r["resume_source"] or "Pending",
            r["resume_file_name"] or "Pending", None, None,
            "None (verified)", "None (verified)", defense_class, why, notes,
        ]
        ws.append(row_values)
        row = ws.max_row
        for c in range(1, len(row_values) + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = BASE_FONT
            cell.border = BORDER
            if c in (3, 21, 22):
                cell.alignment = WRAP
        if r["discovered_at"]:
            dc = ws.cell(row=row, column=10, value=r["discovered_at"].replace(tzinfo=None))
            dc.number_format = "yyyy-mm-dd hh:mm"
            dc.font = BASE_FONT
        if r["url"]:
            uc = ws.cell(row=row, column=7, value=r["url"])
            uc.hyperlink = r["url"]
            uc.font = LINK_FONT
        if r["resume_link"]:
            lc = ws.cell(row=row, column=16, value=r["resume_file_name"])
            lc.hyperlink = r["resume_link"]
            lc.font = LINK_FONT
        if r["draft_created_at"]:
            rc = ws.cell(row=row, column=17, value=r["draft_created_at"].replace(tzinfo=None))
            rc.number_format = "yyyy-mm-dd hh:mm"
            rc.font = BASE_FONT

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.row_dimensions[1].height = 30

    widths = [8, 16, 34, 16, 12, 14, 38, 11, 34, 16, 10, 13, 15, 13, 30, 30, 16, 20, 24, 14, 50, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if len(clean_records) >= 1:
        ref = f"A1:{get_column_letter(len(headers))}{len(clean_records) + 1}"
        table = Table(displayName="NewJobsTable", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
        ws.add_table(table)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\nSaved NEW JOBS sheet ({len(clean_records)} rows) to {OUT_PATH}")
    return clean_records, violations


def _final_report(**k):
    new_records = k["new_records"]
    non_us = sum(1 for r in new_records if classify_location(r["location"]) == "non_us")
    senior = sum(1 for r in new_records if classify_seniority(r["title"], r["description_raw"] or "") == "senior"
                 or _HARD_EXCLUDE_TITLE_PATTERN.search(r["title"]))
    wrong_role = sum(1 for r in new_records if not is_target_role(r["title"]))

    # Independent re-verification of the citizenship/clearance restriction,
    # read straight from the persisted Job rows' own work_auth_status/reason
    # (ground truth) - not a recomputation, not trusted in-memory bookkeeping.
    citizenship_violations = clearance_violations = 0
    if new_records:
        verify_session = get_session_factory()()
        try:
            rows = (
                verify_session.query(Job.id, Job.work_auth_status, Job.work_auth_reason)
                .filter(Job.id.in_([r["id"] for r in new_records]))
                .all()
            )
        finally:
            verify_session.close()
        for _id, status, reason in rows:
            excluded, bucket = citizenship_clearance_exclusion(status, reason)
            if excluded and bucket == "citizenship":
                citizenship_violations += 1
            elif excluded and bucket == "clearance":
                clearance_violations += 1

    seen_canon = {}
    dup_new = 0
    for r in new_records:
        if r["canonical_key"] in seen_canon:
            dup_new += 1
        seen_canon[r["canonical_key"]] = r["id"]
    mapping_mismatches = sum(
        1 for r in new_records
        if r["draft_id"] is not None and not r["resume_link"]
    )

    print("\n" + "=" * 70)
    print("FINAL QUALITY CHECK")
    print("=" * 70)
    print(f"NEW JOBS with non-US location: {non_us} (must be 0)")
    print(f"NEW JOBS with senior/staff/principal/lead/manager/director/head/distinguished: {senior} (must be 0)")
    print(f"NEW JOBS with explicit citizenship requirement: {citizenship_violations} (must be 0)")
    print(f"NEW JOBS with mandatory security clearance: {clearance_violations} (must be 0)")
    print(f"NEW JOBS with wrong role: {wrong_role} (must be 0)")
    print(f"Duplicate NEW JOBS (shared canonical_key within this batch): {dup_new} (must be 0)")
    print(f"Resume/job mapping mismatches (draft exists but no verified PDF link): {mapping_mismatches} (must be 0)")

    gate_failed = any([non_us, senior, citizenship_violations, clearance_violations, wrong_role, dup_new, mapping_mismatches])

    print("\nFINAL VERDICT:", "FAIL - LIVE REFRESH QUALITY ISSUE FOUND" if gate_failed else "PASS - NEW HIGH-QUALITY JOB BATCH READY")


if __name__ == "__main__":
    main()
