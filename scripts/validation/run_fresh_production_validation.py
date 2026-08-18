"""One-off historical run: validated the posted_within_days=7 freshness gate
against live sources. Kept as a record of that run, not meant to be re-run
as-is.

Calls discover_jobs() directly - the real pipeline node, freshness gate
included - then analyze_job/select_master_resume/resume_agent_node on a
bounded, diversity-capped selection from its output. Rebuilds the tracker's
core sheets via build_job_tracker.py, then appends a "NEW JOBS" sheet for
this run.

Makes real Greenhouse/Lever/web-search and Claude API calls.
"""

import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from langgraph.graph import END
from langgraph.types import Send

from src.db.models import Job, SearchConfig
from src.db.session import get_session_factory
from src.discovery.normalize import (
    CURRENTLY_ELIGIBLE,
    POST_DATE_UNKNOWN,
    POSTED_OLD,
    POSTED_RECENTLY,
    classify_freshness,
    classify_location,
    classify_seniority,
    is_target_role,
)
from src.graph.pipeline import analyze_job, discover_jobs, resume_agent_node, select_master_resume
from src.reporting.excel_style import BASE_FONT, BORDER, HEADER_FILL, HEADER_FONT, LINK_FONT, WRAP
from src.reporting.tracker import fetch_records
from src.resume.pdf_render import _pdf_page_count

import httpx
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT_PATH = Path(__file__).resolve().parents[2] / "output" / "job_tracker.xlsx"

MAX_ATS_CALLS = 10
MAX_PER_COMPANY = 2

_DEFENSE_PATTERN = re.compile(
    r"\b(defense|defence|department of defense|dod|national security|intelligence community|"
    r"federal government|public sector|government contract|military|homeland security|"
    r"classified (?:information|program|systems)|intelligence agenc(?:y|ies)|counterterrorism)\b",
    re.IGNORECASE,
)
_ROLE_TIERS = [
    (1, re.compile(r"\bbackend\b.{0,20}\bengineer\b", re.IGNORECASE)),
    (2, re.compile(r"\bsoftware engineer\b|\bsde\s*(?:ii|2)\b", re.IGNORECASE)),
    (3, re.compile(r"\b(?:api|platform|infrastructure)\b.{0,20}\bengineer\b", re.IGNORECASE)),
    (4, re.compile(r"\b(?:machine learning|ml) engineer\b", re.IGNORECASE)),
    (5, re.compile(r"\bai engineer\b|\bapplied ai engineer\b", re.IGNORECASE)),
    (6, re.compile(r"\bdata scientist\b|\banalytics engineer\b", re.IGNORECASE)),
    (7, re.compile(r"\binfrastructure engineer\b|\bplatform engineer\b", re.IGNORECASE)),
]
_EXPERIENCE_RANK = {"entry_level": 0, "0-1": 1, "1-2": 2, "unknown": 3, "2-3": 4, "3-5": 5, "5+": 6}


def role_priority_tier(title: str) -> int:
    for tier, pattern in _ROLE_TIERS:
        if pattern.search(title):
            return tier
    return 9


def classify_defense(title: str, description: str) -> bool:
    return bool(_DEFENSE_PATTERN.search(f"{title}\n{description or ''}"))


def citizenship_clearance_exclusion(work_auth_status: str, work_auth_reason: str | None) -> tuple[bool, str | None]:
    """Reads the ALREADY-COMPUTED work_auth_status/work_auth_reason - no new
    classification logic. production's _split_ineligible() only excludes
    "ineligible"; this additionally excludes a bare "unknown" clearance/
    classified mention - stricter than production's own gate."""
    reason_l = (work_auth_reason or "").lower()
    if work_auth_status == "ineligible":
        bucket = "clearance" if "clearance" in reason_l else "citizenship"
        return True, bucket
    if work_auth_status == "unknown" and any(m in reason_l for m in ("clearance", "classified", "us person")):
        bucket = "clearance" if "clearance" in reason_l else "citizenship"
        return True, bucket
    return False, None


def main() -> None:
    session = get_session_factory()()
    try:
        baseline_max_job_id = session.query(Job.id).order_by(Job.id.desc()).first()
        baseline_max_job_id = baseline_max_job_id[0] if baseline_max_job_id else 0
        config = session.query(SearchConfig).first()
        posted_within_days = config.posted_within_days
    finally:
        session.close()

    print(f"Baseline max job id (pre-run): {baseline_max_job_id}")
    print(f"Active freshness gate: posted_within_days={posted_within_days}\n")
    assert posted_within_days == 7, f"expected posted_within_days=7, got {posted_within_days}"

    print("=== LIVE DISCOVERY (real discover_jobs() - Greenhouse + Lever + web-search) ===")
    result = discover_jobs({})
    discovery_stats = result["discovery_stats"]
    new_jobs = result["new_jobs"]  # already: genuinely new + CURRENTLY_ELIGIBLE + POSTED_RECENTLY + not work-auth-"ineligible"

    print(f"raw_fetched: {discovery_stats.get('raw_fetched', 'n/a - see per-stage stats below')}")
    for k in ("raw_fetched", "non_us_rejected", "non_full_time_rejected", "title_rejected", "senior_rejected", "eligible"):
        if k in discovery_stats:
            print(f"  {k}: {discovery_stats[k]}")
    print(f"duplicates_removed: {discovery_stats.get('duplicates_removed', 0)} "
          f"(cross_source: {discovery_stats.get('cross_source_duplicates_removed', 0)})")
    print(f"new_jobs_selected (persisted this run): {discovery_stats.get('new_jobs_selected', 0)}")
    print(f"ineligible_or_ambiguous_excluded (current_eligibility gate): {discovery_stats.get('ineligible_or_ambiguous_excluded', 0)}")
    print(f"stale_or_unknown_date_excluded (freshness gate): {discovery_stats.get('stale_or_unknown_date_excluded', 0)}")
    print(f"immigration_ineligible (explicit citizenship/PR/clearance-tied-to-citizenship): {discovery_stats.get('immigration_ineligible', 0)}")
    print(f"pending_resume_jobs (unrelated crash-recovery queue, not touched this run): {discovery_stats.get('pending_resume_jobs', 0)}")

    # Granular recent-vs-old-vs-unknown breakdown, read back from the
    # newly-persisted Job rows themselves (independent re-derivation, not
    # trusted in-memory discovery_stats bookkeeping).
    session = get_session_factory()()
    try:
        now = datetime.now(timezone.utc)
        new_rows = session.query(Job).filter(Job.id > baseline_max_job_id).all()
        recent_ct = sum(1 for j in new_rows if classify_freshness(j.posted_at, posted_within_days, now) == POSTED_RECENTLY)
        old_ct = sum(1 for j in new_rows if classify_freshness(j.posted_at, posted_within_days, now) == POSTED_OLD)
        unknown_ct = sum(1 for j in new_rows if classify_freshness(j.posted_at, posted_within_days, now) == POST_DATE_UNKNOWN)
        print(f"\nPersisted-this-run breakdown (independent re-check): {len(new_rows)} total | "
              f"POSTED_RECENTLY={recent_ct} POSTED_OLD={old_ct} POST_DATE_UNKNOWN={unknown_ct}")
    finally:
        session.close()

    print(f"\ngenuinely new + CURRENTLY_ELIGIBLE + POSTED_RECENTLY (candidate pool for this run): {len(new_jobs)}")

    if not new_jobs:
        print("\nNo genuinely-new, fresh, eligible jobs this run - nothing to process.")
        _refresh_excel([], baseline_max_job_id, {})
        print("\nFINAL VERDICT: FAIL - LIVE PIPELINE ISSUE FOUND (no fresh eligible jobs this run - see report for root cause)")
        return

    # Additional citizenship/clearance stricter bar (see module docstring) +
    # defense classification (reporting/priority only) + priority ranking.
    citizenship_excluded = clearance_excluded = 0
    survivors = []
    for job in new_jobs:
        excluded, bucket = citizenship_clearance_exclusion(job.get("work_auth_status"), job.get("work_auth_reason"))
        if excluded:
            if bucket == "citizenship":
                citizenship_excluded += 1
            else:
                clearance_excluded += 1
            continue
        survivors.append(job)

    print(f"Additional citizenship-excluded (this batch's stricter bar): {citizenship_excluded}")
    print(f"Additional clearance-excluded (this batch's stricter bar): {clearance_excluded}")
    print(f"Survivors after stricter citizenship/clearance bar: {len(survivors)}")

    def sort_key(job):
        tier = role_priority_tier(job["title"])
        defense_rank = 1 if classify_defense(job["title"], job["description_raw"]) else 0
        exp_rank = _EXPERIENCE_RANK.get(classify_seniority(job["title"], job["description_raw"] or ""), 3)
        return (tier, defense_rank, exp_rank, job["company"], job["title"])

    survivors_sorted = sorted(survivors, key=sort_key)

    selected = []
    company_counts: dict[str, int] = {}
    for job in survivors_sorted:
        key = job["company"].strip().lower()
        if company_counts.get(key, 0) >= MAX_PER_COMPANY:
            continue
        selected.append(job)
        company_counts[key] = company_counts.get(key, 0) + 1
        if len(selected) >= MAX_ATS_CALLS:
            break

    print(f"\nSelected for ATS (bounded <= {MAX_ATS_CALLS}, priority + diversity ranked): {len(selected)}")
    print(f"Companies represented: {sorted(company_counts.keys())}")
    assert len(selected) <= MAX_ATS_CALLS

    # === ATS + Resume Agent - REAL calls, reusing analyze_job/select_master_resume/resume_agent_node unmodified ===
    print("\n=== ATS (real Claude calls via src.graph.pipeline.analyze_job, unmodified) ===")
    ats_calls_made = 0
    resumes_master_only = 0
    resumes_generated = 0
    ats_records = []
    for job in selected:
        if ats_calls_made >= MAX_ATS_CALLS:
            break
        state = {
            "job": job,
            "profile": _profile_json(),
            "reject_threshold": result["reject_threshold"],
            "no_tailor_threshold": result["no_tailor_threshold"],
            "master_raw_latex": _master_raw_latex(),
            "github_context": _github_context(_master_raw_latex()),
        }
        command = analyze_job(state)
        ats_calls_made += 1
        results = command.update.get("results", [])
        if not results:
            print(f"  job {job['id']} ({job['company']} - {job['title']}): ATS call failed, left pending.")
            continue
        r = results[0]
        ats_records.append({"id": job["id"], "company": job["company"], "title": job["title"], "score": r["match_score"], "status": r["status"]})
        print(f"  job {job['id']} ({job['company']} - {job['title']}): score={r['match_score']} status={r['status']}")

        if command.goto == END:
            continue
        if isinstance(command.goto, Send):
            if command.goto.node == "select_master_resume":
                select_master_resume(command.goto.arg)
                resumes_master_only += 1
            elif command.goto.node == "resume_agent":
                resume_agent_node(command.goto.arg)
                resumes_generated += 1

    print(f"\nATS calls made: {ats_calls_made} | passed->master: {resumes_master_only} | passed->resume_agent (tailored attempt): {resumes_generated}")

    print("\n=== Independent re-verification against DB + filesystem ===")
    records, db_counts, _pwd = fetch_records()
    # Scoped by "actually ATS-processed this run" (selected_ids), not "row
    # id newer than baseline" - a re-surfaced previously-pending job has an
    # OLD id but belongs on this run's NEW JOBS sheet; a newly-inserted but
    # not-selected row should not.
    selected_ids = {j["id"] for j in selected}
    new_records = [r for r in records if r["id"] in selected_ids]
    print(f"DB counts now: {db_counts}")
    print(f"Jobs processed by ATS this run (re-read from DB): {len(new_records)}")

    print("\n=== Rebuilding core Excel sheets (build_job_tracker.py, unmodified) then appending NEW JOBS ===")
    tracker_script = Path(__file__).resolve().parents[1] / "reporting" / "build_job_tracker.py"
    subprocess.run([sys.executable, str(tracker_script)], check=True)

    _refresh_excel(new_records, baseline_max_job_id, discovery_stats)
    _final_report(
        discovery_stats=discovery_stats, recent_ct=recent_ct, old_ct=old_ct, unknown_ct=unknown_ct,
        new_jobs_pool=len(new_jobs), citizenship_excluded=citizenship_excluded, clearance_excluded=clearance_excluded,
        selected=selected, ats_calls_made=ats_calls_made, resumes_master_only=resumes_master_only,
        resumes_generated=resumes_generated, new_records=new_records, baseline_max_job_id=baseline_max_job_id,
        db_counts=db_counts, posted_within_days=posted_within_days, ats_records=ats_records,
    )


_PROFILE_CACHE = {}


def _profile_row():
    if "row" not in _PROFILE_CACHE:
        from src.db.models import CandidateProfile
        session = get_session_factory()()
        try:
            row = session.query(CandidateProfile).first()
            if row is None:
                raise RuntimeError("candidate_profile is empty")
            _PROFILE_CACHE["row"] = (row.structured_json, row.raw_latex)
        finally:
            session.close()
    return _PROFILE_CACHE["row"]


def _profile_json():
    return _profile_row()[0]


def _master_raw_latex():
    return _profile_row()[1]


_GH_CACHE = {}


def _github_context(master_raw_latex):
    if "ctx" not in _GH_CACHE:
        from src.integrations.github import fetch_github_context
        _GH_CACHE["ctx"] = fetch_github_context(master_raw_latex)
    return _GH_CACHE["ctx"]


def _quality_check(records):
    clean, violations = [], []
    for r in records:
        loc = classify_location(r["location"])
        role_ok = is_target_role(r["title"])
        sen = classify_seniority(r["title"], r["description_raw"] or "")
        problems = []
        if loc == "non_us":
            problems.append("non-US location")
        if sen == "senior":
            problems.append("senior/staff/principal/lead title")
        if not role_ok:
            problems.append("wrong role family")
        if r["current_eligibility"] != CURRENTLY_ELIGIBLE:
            problems.append(f"current_eligibility={r['current_eligibility']}")
        if not r["canonical_key"]:
            problems.append("invalid canonical_key")
        if problems:
            violations.append((r["id"], problems))
        else:
            clean.append(r)
    return clean, violations


def _refresh_excel(new_records, baseline_max_job_id, discovery_stats):
    clean_records, violations = _quality_check(new_records) if new_records else ([], [])
    if violations:
        print(f"\nQUALITY CHECK: removed {len(violations)} row(s) that failed re-verification: {violations}")

    # Fetch posted_at/freshness/work_auth per remaining record for the sheet
    # (fetch_records() doesn't expose work_auth_status - query it directly).
    ids = [r["id"] for r in clean_records]
    work_auth_by_id = {}
    if ids:
        session = get_session_factory()()
        try:
            rows = session.query(Job.id, Job.work_auth_status, Job.work_auth_reason).filter(Job.id.in_(ids)).all()
            work_auth_by_id = {i: (s, rsn) for i, s, rsn in rows}
        finally:
            session.close()

    clean_records.sort(key=lambda r: (-(r["ats_score"] or -1), r["id"]))

    if not OUT_PATH.exists():
        raise RuntimeError(f"{OUT_PATH} does not exist - build_job_tracker.py must run first")
    wb = openpyxl.load_workbook(OUT_PATH)
    if "NEW JOBS" in wb.sheetnames:
        del wb["NEW JOBS"]
    ws = wb.create_sheet("NEW JOBS")

    headers = [
        "Job ID", "Company", "Title", "Location", "Posted Date", "Days Since Posted",
        "ATS Score", "ATS Status", "Job URL", "Resume Status", "Resume PDF Filename",
        "Resume PDF Link", "Discovery State", "Notes",
    ]
    ws.append(headers)

    for rec in clean_records:
        # "Pending" must mean genuinely pending (still could get a resume
        # later) - an ATS-rejected/ineligible job will NEVER get one, so
        # labeling it "Pending" would be misleading, not just incomplete.
        if rec["status"] in ("RESUME NOT REQUIRED", "RESUME DECLINED", "RESUME COMPLETED"):
            resume_status = rec["status"]
        elif rec["ats_status"] == "rejected":
            resume_status = "N/A (ATS Rejected)"
        elif rec["ats_status"] == "ineligible":
            resume_status = "N/A (Ineligible)"
        else:
            resume_status = "Pending"
        discovery_state = "NEW THIS RUN" if rec["id"] > baseline_max_job_id else "EXISTING"
        _status, reason = work_auth_by_id.get(rec["id"], (None, None))
        notes_parts = []
        if rec["eligibility_reasons"]:
            notes_parts.append("; ".join(rec["eligibility_reasons"]))
        if not notes_parts:
            notes_parts.append("No known rule violations")
        pdf_filename_display = rec["resume_file_name"] or ("N/A" if resume_status.startswith("N/A") else "Pending")
        row_values = [
            rec["id"], rec["company"], rec["title"], rec["location"],
            None, rec["days_since_posted"] if rec["days_since_posted"] is not None else "Unknown",
            rec["ats_score"], rec["ats_status"] or "Pending", None, resume_status,
            pdf_filename_display, None, discovery_state, "; ".join(notes_parts),
        ]
        ws.append(row_values)
        row = ws.max_row
        for c in range(1, len(row_values) + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = BASE_FONT
            cell.border = BORDER
            if c in (3, 14):
                cell.alignment = WRAP
        if rec["posted_at"]:
            dc = ws.cell(row=row, column=5, value=rec["posted_at"].replace(tzinfo=None))
            dc.number_format = "yyyy-mm-dd hh:mm"
            dc.font = BASE_FONT
        else:
            ws.cell(row=row, column=5, value="Unknown").font = BASE_FONT
        if rec["url"]:
            uc = ws.cell(row=row, column=9, value=rec["url"])
            uc.hyperlink = rec["url"]
            uc.font = LINK_FONT
        if rec["resume_link"]:
            lc = ws.cell(row=row, column=12, value=rec["resume_file_name"])
            lc.hyperlink = rec["resume_link"]
            lc.font = LINK_FONT

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.row_dimensions[1].height = 26

    widths = [8, 16, 34, 16, 16, 12, 10, 13, 38, 15, 30, 30, 14, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if len(clean_records) >= 1:
        ref = f"A1:{get_column_letter(len(headers))}{len(clean_records) + 1}"
        table = Table(displayName="NewJobsTable", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
        ws.add_table(table)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\nSaved NEW JOBS sheet ({len(clean_records)} rows) to {OUT_PATH}")
    return clean_records, violations, work_auth_by_id


def _final_report(**k):
    new_records = k["new_records"]
    posted_within_days = k["posted_within_days"]
    now = datetime.now(timezone.utc)

    not_within_7 = sum(
        1 for r in new_records
        if classify_freshness(r["posted_at"], posted_within_days, now) != POSTED_RECENTLY
    )
    non_us = sum(1 for r in new_records if classify_location(r["location"]) == "non_us")
    senior = sum(1 for r in new_records if classify_seniority(r["title"], r["description_raw"] or "") == "senior")

    ids = [r["id"] for r in new_records]
    citizenship_violations = clearance_violations = 0
    if ids:
        session = get_session_factory()()
        try:
            rows = session.query(Job.id, Job.work_auth_status, Job.work_auth_reason).filter(Job.id.in_(ids)).all()
        finally:
            session.close()
        for _id, status, reason in rows:
            excluded, bucket = citizenship_clearance_exclusion(status, reason)
            if excluded and bucket == "citizenship":
                citizenship_violations += 1
            elif excluded and bucket == "clearance":
                clearance_violations += 1

    seen_canon = {}
    dup_new = 0
    for r in new_records:
        if r["canonical_key"] in seen_canon:
            dup_new += 1
        seen_canon[r["canonical_key"]] = r["id"]

    # PDF existence / one-page / job-mapping / ATS-score-matches-DB checks.
    pdf_missing, pdf_multipage, pdf_ok, ats_mismatch = [], [], [], []
    for r in new_records:
        if r["draft_id"] is not None:
            if not r["resume_link"]:
                pdf_missing.append(r["id"])
            else:
                p = Path(r["resume_file_path"])
                if not p.exists():
                    pdf_missing.append(r["id"])
                else:
                    try:
                        pages = _pdf_page_count(p)
                    except Exception:
                        pdf_multipage.append((r["id"], "unreadable"))
                        continue
                    if pages != 1:
                        pdf_multipage.append((r["id"], pages))
                    else:
                        pdf_ok.append(r["id"])

    # ATS score DB-match check: cross-reference against the live analyze_job
    # results captured during this run.
    ats_by_id = {a["id"]: a["score"] for a in k["ats_records"]}
    for r in new_records:
        if r["id"] in ats_by_id and r["ats_score"] != ats_by_id[r["id"]]:
            ats_mismatch.append(r["id"])

    # Job URL reachability - lightweight live check, this run's NEW JOBS only.
    url_failures = []
    if new_records:
        with httpx.Client(timeout=15, follow_redirects=True) as client:
            for r in new_records:
                if not r["url"]:
                    url_failures.append((r["id"], "empty URL"))
                    continue
                try:
                    resp = client.head(r["url"])
                    if resp.status_code >= 400:
                        resp = client.get(r["url"])
                    if resp.status_code >= 400:
                        url_failures.append((r["id"], resp.status_code))
                except httpx.HTTPError as exc:
                    url_failures.append((r["id"], str(exc)[:100]))

    print("\n" + "=" * 70)
    print("FINAL REPORT")
    print("=" * 70)
    ds = k["discovery_stats"]
    print(f"Raw jobs fetched: {ds.get('raw_fetched', 'n/a')}")
    print(f"Fresh jobs found (POSTED_RECENTLY, persisted this run): {k['recent_ct']}")
    print(f"  (POSTED_OLD: {k['old_ct']}, POST_DATE_UNKNOWN: {k['unknown_ct']}, persisted this run)")
    print(f"Genuinely new + CURRENTLY_ELIGIBLE + POSTED_RECENTLY pool: {k['new_jobs_pool']}")
    print(f"Additional citizenship-excluded (this batch): {k['citizenship_excluded']}")
    print(f"Additional clearance-excluded (this batch): {k['clearance_excluded']}")
    print(f"Jobs selected/processed by ATS: {len(k['selected'])} (ATS calls made: {k['ats_calls_made']})")
    print(f"Resumes generated (tailored attempt): {k['resumes_generated']} | master-only: {k['resumes_master_only']}")
    print(f"PDFs verified one-page/OK: {len(pdf_ok)} | missing: {len(pdf_missing)} | multipage/unreadable: {len(pdf_multipage)}")
    print(f"Final NEW JOBS count: {len(new_records)}")
    print(f"Companies represented in NEW JOBS: {sorted({r['company'] for r in new_records})}")
    print(f"Excel output path: {OUT_PATH}")
    for r in new_records:
        if r["resume_file_path"]:
            print(f"  PDF: job {r['id']} ({r['company']}) -> {r['resume_file_path']}")

    print("\n" + "-" * 70)
    print("FINAL VERIFICATION")
    print("-" * 70)
    print(f"NEW JOBS not posted within {posted_within_days} days: {not_within_7} (must be 0)")
    print(f"NEW JOBS with non-US location: {non_us} (must be 0)")
    print(f"NEW JOBS with senior/staff/principal/lead title: {senior} (must be 0)")
    print(f"NEW JOBS with citizenship requirement: {citizenship_violations} (must be 0)")
    print(f"NEW JOBS with mandatory/ambiguous clearance mention: {clearance_violations} (must be 0)")
    print(f"Duplicate NEW JOBS (shared canonical_key): {dup_new} (must be 0)")
    print(f"Job URL failures (live HEAD/GET check): {len(url_failures)} (must be 0) {url_failures if url_failures else ''}")
    print(f"PDFs missing: {len(pdf_missing)} (must be 0) {pdf_missing if pdf_missing else ''}")
    print(f"PDFs not exactly one page: {len(pdf_multipage)} (must be 0) {pdf_multipage if pdf_multipage else ''}")
    print(f"ATS score DB mismatches: {len(ats_mismatch)} (must be 0) {ats_mismatch if ats_mismatch else ''}")

    gate_failed = any([
        not_within_7, non_us, senior, citizenship_violations, clearance_violations,
        dup_new, url_failures, pdf_missing, pdf_multipage, ats_mismatch,
    ])
    print("\nFINAL VERDICT:", "FAIL - LIVE PIPELINE ISSUE FOUND" if gate_failed else "PASS - FRESH LIVE PIPELINE VERIFIED")


if __name__ == "__main__":
    main()
