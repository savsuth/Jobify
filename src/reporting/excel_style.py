"""Shared openpyxl styling for the job tracker workbook and its one-off
validation sheets. Kept in one place because build_job_tracker.py and the
scripts under scripts/validation/ all render the same visual style.
"""

from datetime import datetime

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT_NAME = "Arial"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BASE_FONT = Font(name=FONT_NAME, size=10)
LINK_FONT = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DATE_FMT = "yyyy-mm-dd hh:mm"

ELIGIBILITY_COLORS = {
    "CURRENTLY_ELIGIBLE": "C9E7B7",
    "CURRENTLY_INELIGIBLE": "F4CCCC",
    "AMBIGUOUS_REQUIRES_REVIEW": "FFF2CC",
}
STATUS_COLORS = {
    "ATS PENDING": "FFF2CC", "ATS REJECTED": "F4CCCC", "ATS INELIGIBLE": "D9D2E9",
    "ATS PASSED — RESUME PENDING": "FCE5CD", "RESUME NOT REQUIRED": "D9EAD3",
    "RESUME DECLINED": "EAD1DC", "RESUME COMPLETED": "C9E7B7",
}


def as_naive_dt(dt):
    return dt.replace(tzinfo=None) if dt else None


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_table(ws, name, ncols, nrows):
    if nrows < 1:
        return
    ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
    ws.add_table(table)


def write_row(ws, values, wrap_cols=(), date_cols=(), fill_key=None, fill_col=None):
    ws.append(values)
    r = ws.max_row
    for c in range(1, len(values) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = BASE_FONT
        cell.border = BORDER
        if c in wrap_cols:
            cell.alignment = WRAP
        if c in date_cols and isinstance(values[c - 1], datetime):
            cell.number_format = DATE_FMT
    if fill_key and fill_col:
        color = ELIGIBILITY_COLORS.get(fill_key) or STATUS_COLORS.get(fill_key)
        if color:
            ws.cell(row=r, column=fill_col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return r


def add_hyperlink(ws, row, col, url, label=None):
    cell = ws.cell(row=row, column=col)
    if label is not None:
        cell.value = label
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT
